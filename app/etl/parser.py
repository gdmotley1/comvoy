"""
Parse the Comvoy Multi-Brand Excel report into structured dicts.

Sheets parsed:
  - Executive Summary → report metadata (date, totals)
  - All Dealers → dealer list with summary stats
  - Dealer Brand Matrix → dealer × brand vehicle counts
  - Dealer Body Type Matrix → dealer × body type vehicle counts
  - Smyrna Product Analysis → Smyrna-specific dealer details
  - Brand Summary → brand lookup list
  - Body Type Summary → body type lookup list
"""

import re
from datetime import date, datetime
from pathlib import Path

import openpyxl


def parse_report(file_path: str | Path) -> dict:
    """Parse the full Excel report and return structured data."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    result = {
        "metadata": _parse_metadata(wb["Executive Summary"]),
        "brands": _parse_brands(wb["Brand Summary"]),
        "body_types": _parse_body_types(wb["Body Type Summary"]),
        "dealers": _parse_all_dealers(wb["All Dealers"]),
        "brand_matrix": _parse_brand_matrix(wb["Dealer Brand Matrix"]),
        "body_type_matrix": _parse_body_type_matrix(wb["Dealer Body Type Matrix"]),
        "smyrna_details": _parse_smyrna(wb["Smyrna Product Analysis"]),
    }
    wb.close()
    return result


def _parse_metadata(ws) -> dict:
    """Extract report date and territory totals from Executive Summary."""
    # Row 2 contains the date string like "12-State Territory  |  Prepared February 18, 2026  ..."
    date_text = str(ws.cell(row=2, column=1).value or "")
    report_date = _extract_date(date_text)

    # Row 5 has the headline numbers: 588, 13439, 13, 25
    return {
        "report_date": report_date,
        "total_dealers": _int(ws.cell(row=5, column=1).value),
        "total_vehicles": _int(ws.cell(row=5, column=3).value),
        "total_brands": _int(ws.cell(row=5, column=5).value),
        "total_body_types": _int(ws.cell(row=5, column=7).value),
    }


def _parse_brands(ws) -> list[dict]:
    """Parse Brand Summary sheet into brand list."""
    brands = []
    # Header row is 4, data starts at 5
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        name = row[0]
        if not name or str(name).strip().upper() == "TOTAL":
            break
        brands.append({"name": str(name).strip()})
    return brands


def _parse_body_types(ws) -> list[dict]:
    """Parse Body Type Summary sheet into body type list."""
    body_types = []
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        name = row[0]
        if not name or str(name).strip().upper() == "TOTAL":
            break
        body_types.append({"name": str(name).strip()})
    return body_types


def _parse_all_dealers(ws) -> list[dict]:
    """Parse All Dealers sheet (588 rows)."""
    dealers = []
    # Headers at row 4: Rank, Dealer Name, City, ST, Total Vehicles, Brands, Body Types,
    #                    Top Brand, Smyrna Units, Smyrna %, Top Body Types
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        rank = row[0]
        name = row[1]
        if not name or not rank:
            continue
        dealers.append({
            "rank": _int(rank),
            "name": str(name).strip(),
            "city": str(row[2] or "").strip(),
            "state": str(row[3] or "").strip().upper(),
            "total_vehicles": _int(row[4]),
            "brand_count": _int(row[5]),
            "body_type_count": _int(row[6]),
            "top_brand": str(row[7] or "").strip(),
            "smyrna_units": _int(row[8]),
            "smyrna_percentage": _float(row[9]),
            "top_body_types": str(row[10] or "").strip(),
        })
    return dealers


def _parse_brand_matrix(ws) -> list[dict]:
    """Parse Dealer Brand Matrix (dealer × 13 brands).

    Columns: Dealer Name, City, ST, Brands, Ford, Chevy, GMC, Ram, Frtlnr,
             W.Star, Peter, Knwrth, Intl, Volvo, Mack, Isuzu, Hino, Total
    """
    # Read header row to get brand column names
    header_row = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]

    # Map short column names to full brand names
    brand_col_map = {
        "Ford": "Ford",
        "Chevy": "Chevrolet",
        "GMC": "GMC",
        "Ram": "Ram",
        "Frtlnr": "Freightliner",
        "W.Star": "Western Star",
        "Peter": "Peterbilt",
        "Knwrth": "Kenworth",
        "Intl": "International",
        "Volvo": "Volvo",
        "Mack": "Mack",
        "Isuzu": "Isuzu",
        "Hino": "Hino",
    }

    # Find which columns map to which brands
    brand_columns = {}  # col_index -> brand_name
    for idx, header in enumerate(header_row):
        header_str = str(header or "").strip()
        if header_str in brand_col_map:
            brand_columns[idx] = brand_col_map[header_str]

    rows = []
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        dealer_name = row[0]
        if not dealer_name:
            continue
        dealer_name = str(dealer_name).strip()
        city = str(row[1] or "").strip()
        state = str(row[2] or "").strip().upper()

        for col_idx, brand_name in brand_columns.items():
            count = _int(row[col_idx])
            if count and count > 0:
                rows.append({
                    "dealer_name": dealer_name,
                    "city": city,
                    "state": state,
                    "brand": brand_name,
                    "vehicle_count": count,
                })
    return rows


def _parse_body_type_matrix(ws) -> list[dict]:
    """Parse Dealer Body Type Matrix (dealer × 25 body types).

    Columns: Dealer Name, City, ST, Types, then 25 body type columns, then Total
    """
    header_row = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]

    # Map short column headers to full body type names
    bt_col_map = {
        "Svc\nTruck": "Service Trucks",
        "Box\nTruck": "Box Trucks",
        "Box\nVan": "Box Vans",
        "Flatbed": "Flatbed Trucks",
        "Svc Util\nVan": "Service Utility Vans",
        "Refrig": "Refrigerated",
        "Dump": "Dump Trucks",
        "Cutaway": "Cutaways",
        "Lndscpe\nDump": "Landscape Dumps",
        "Mechanic": "Mechanic Body",
        "Stake\nBed": "Stake Beds",
        "Contract": "Contractor Trucks",
        "Encl Svc": "Enclosed Service",
        "Rollback": "Rollback",
        "Combo": "Combo Body",
        "Bucket": "Bucket Trucks",
        "Chipper": "Chipper Trucks",
        "Dovetail": "Dovetail Landscapes",
        "Hauler": "Hauler Body",
        "Hooklift": "Hooklift",
        "Crane": "Crane Trucks",
        "Wrecker": "Wrecker",
        "Flat\nDump": "Flatbed Dump",
        "Welder": "Welder Body",
        "Roll-Off": "Roll-Off",
    }

    body_type_columns = {}
    for idx, header in enumerate(header_row):
        header_str = str(header or "").strip()
        if header_str in bt_col_map:
            body_type_columns[idx] = bt_col_map[header_str]

    rows = []
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        dealer_name = row[0]
        if not dealer_name:
            continue
        dealer_name = str(dealer_name).strip()
        city = str(row[1] or "").strip()
        state = str(row[2] or "").strip().upper()

        for col_idx, bt_name in body_type_columns.items():
            count = _int(row[col_idx])
            if count and count > 0:
                rows.append({
                    "dealer_name": dealer_name,
                    "city": city,
                    "state": state,
                    "body_type": bt_name,
                    "vehicle_count": count,
                })
    return rows


def _parse_smyrna(ws) -> list[dict]:
    """Parse Smyrna Product Analysis sheet (24 dealers)."""
    details = []
    # Header at row 4: Rank, Dealer Name, City, ST, Smyrna Units, Dealer Total,
    #                   Smyrna %, Top Smyrna Body Types, Avg Days Since Upfit
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        rank = row[0]
        name = row[1]
        if not name or str(name).strip().upper() == "TOTAL":
            break
        if not rank:
            continue
        details.append({
            "dealer_name": str(name).strip(),
            "city": str(row[2] or "").strip(),
            "state": str(row[3] or "").strip().upper(),
            "smyrna_units": _int(row[4]),
            "dealer_total": _int(row[5]),
            "smyrna_percentage": _float(row[6]),
            "top_smyrna_body_types": str(row[7] or "").strip(),
            "avg_days_since_upfit": _int(row[8]),
        })
    return details


# --- Helpers ---

def _int(val) -> int:
    if val is None:
        return 0
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return 0


def _float(val) -> float:
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _extract_date(text: str) -> date:
    """Extract date from text like 'Prepared February 18, 2026'."""
    match = re.search(r"Prepared\s+(\w+\s+\d{1,2},?\s+\d{4})", text)
    if match:
        date_str = match.group(1).replace(",", "")
        try:
            return datetime.strptime(date_str, "%B %d %Y").date()
        except ValueError:
            pass
    # Fallback: try to find any date-like pattern
    match = re.search(r"(\w+)\s+(\d{1,2}),?\s+(\d{4})", text)
    if match:
        try:
            return datetime.strptime(f"{match.group(1)} {match.group(2)} {match.group(3)}", "%B %d %Y").date()
        except ValueError:
            pass
    return date.today()
