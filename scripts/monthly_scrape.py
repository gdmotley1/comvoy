#!/usr/bin/env python3
"""
Comvoy Monthly Scrape — Master Script
======================================
Single command to run the full monthly intelligence cycle:

  python scripts/monthly_scrape.py
  python scripts/monthly_scrape.py --resume   (continue interrupted scrape)

Steps:
  1. Scrape Comvoy inventory (13 brands x 25 bodies x 12 states)
  2. Scrape Smyrna/WTS inventory (VIN source of truth)
  3. Diff against previous month's scrape
  4. Build 9-sheet Excel report with full formatting
  5. Print validation summary

Outputs (in scrape_output/):
  inventory_YYYY-MM-DD.csv
  smyrna_inventory_YYYY-MM-DD.csv
  new_vehicles_YYYY-MM-DD.csv
  sold_vehicles_YYYY-MM-DD.csv
  price_changes_YYYY-MM-DD.csv
  Comvoy_Multi_Brand_Report_YYYY-MM-DD.xlsx
  monthly_scrape_log_YYYY-MM-DD.txt
"""

import sys
import io
import os
import re
import csv
import json
import time
import math
import argparse
import requests
from datetime import datetime
from collections import OrderedDict, Counter, defaultdict
from bs4 import BeautifulSoup

# openpyxl for Excel report
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# ═══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════════

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
}

DELAY = 1.0             # seconds between Comvoy requests
SMYRNA_DELAY = 0.5      # seconds between Smyrna WTS requests
PER_PAGE = 30            # Comvoy results per page
SMYRNA_PER_PAGE = 10     # Smyrna WTS results per page
SMYRNA_MAX_PAGES = 20    # safety limit for Smyrna pagination
MAX_EMPTY = 3            # consecutive empty JSON-LD pages before stopping
REQUEST_TIMEOUT = 30
PROGRESS_SAVE_EVERY = 10 # save progress every N combos

# ── Excluded Dealers ─────────────────────────────────────────────────────────
# Dealers to exclude from scrape results (rental/national chains, not prospects)
EXCLUDED_DEALER_PATTERNS = [
    'penske',
    'mhc ',
    'ryder',
]

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(BASE_DIR, '..', 'scrape_output')

# ── Territory ────────────────────────────────────────────────────────────────

STATES = ['AL', 'AR', 'FL', 'GA', 'KY', 'LA', 'MS', 'NC', 'OK', 'SC', 'TN', 'TX']

BRANDS = {
    'Ford': 'Ford',
    'Chevrolet': 'Chevrolet',
    'GMC': 'GMC',
    'Ram': 'Ram',
    'Freightliner': 'Freightliner',
    'Western+Star': 'Western Star',
    'Peterbilt': 'Peterbilt',
    'Kenworth': 'Kenworth',
    'International': 'International',
    'Volvo': 'Volvo',
    'Mack': 'Mack',
    'Isuzu': 'Isuzu',
    'Hino': 'Hino',
}

BODY_TYPES = {
    'box-trucks-1al0': 'Box Trucks',
    'box-vans-18fs': 'Box Vans',
    'bucket-trucks-464k': 'Bucket Trucks',
    'chipper-trucks-17cg': 'Chipper Trucks',
    'combo-trucks-17k0': 'Combo Trucks',
    'contractor-trucks-17or': 'Contractor Trucks',
    'crane-trucks-189l': 'Crane Trucks',
    'cutaway-vans-18eh': 'Cutaways',
    'landscape-trucks-1a96': 'Dovetail Landscapes',
    'dump-trucks-1das': 'Dump Trucks',
    'enclosed-service-trucks-5jzo': 'Enclosed Service',
    'flatbed-dump-trucks-5fvx': 'Flatbed Dump',
    'flatbed-trucks-1fcd': 'Flatbed Trucks',
    'hauler-trucks-1dpb': 'Hauler Body',
    'hooklift-trucks-1dwy': 'Hooklift',
    'landscape-dump-trucks-1e7a': 'Landscape Dumps',
    'mechanic-trucks-1eww': 'Mechanic Body',
    'refrigerated-trucks-1gta': 'Refrigerated',
    'roll-off-trucks-1hee': 'Roll-Off',
    'rollback-trucks-1hgc': 'Rollback',
    'service-trucks-1hmi': 'Service Trucks',
    'stake-bed-trucks-1lk7': 'Stake Beds',
    'utility-vans-1jlo': 'Service Utility Vans',
    'welding-trucks-1n67': 'Welder',
    'tow-trucks-1nbp': 'Wrecker',
}

CSV_FIELDS = [
    'vin', 'listing_id', 'condition', 'year', 'brand', 'model',
    'body_type', 'body_builder', 'price', 'transmission', 'fuel_type',
    'color', 'dealer_name', 'city', 'state', 'listing_url', 'image_url',
    'scrape_date',
]

SMYRNA_CSV_FIELDS = ['vin', 'name', 'price', 'brand', 'model', 'builder']

# ── Report Styling ───────────────────────────────────────────────────────────

NAVY = "001F4E79"
WHITE = "00FFFFFF"
GRAY = "00808080"
BROWN = "008B4513"
LIGHT_BLUE = "00B4C6E7"
ZEBRA_BLUE = "00F2F7FB"
GOLD = "00FFF2CC"
GREEN = "00C6EFCE"
BLUE_SUMMARY = "00BDD7EE"
PEACH = "00FFF8F0"
LIGHT_RED = "00FFF0F0"
LIGHT_YELLOW = "00FFFDE7"
LIGHT_GRAY = "00F0F0F0"

navy_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
zebra_fill = PatternFill(start_color=ZEBRA_BLUE, end_color=ZEBRA_BLUE, fill_type="solid")
gold_fill = PatternFill(start_color=GOLD, end_color=GOLD, fill_type="solid")
green_fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
blue_fill = PatternFill(start_color=BLUE_SUMMARY, end_color=BLUE_SUMMARY, fill_type="solid")
peach_fill = PatternFill(start_color=PEACH, end_color=PEACH, fill_type="solid")
light_red_fill = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid")
light_yellow_fill = PatternFill(start_color=LIGHT_YELLOW, end_color=LIGHT_YELLOW, fill_type="solid")
light_gray_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")

thin_border_side = Side(style="thin", color=LIGHT_BLUE)
thin_border = Border(left=thin_border_side, right=thin_border_side,
                     top=thin_border_side, bottom=thin_border_side)

center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
center = Alignment(horizontal="center", vertical="center")
left_center = Alignment(horizontal="left", vertical="center")
left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

BRAND_ORDER = ["Ford", "Chevrolet", "GMC", "Ram", "Freightliner", "Western Star",
               "Peterbilt", "Kenworth", "International", "Volvo", "Mack", "Isuzu", "Hino"]

BRAND_ABBREV = {"Ford": "Ford", "Chevrolet": "Chevy", "GMC": "GMC", "Ram": "Ram",
                "Freightliner": "Frtlnr", "Western Star": "W.Star", "Peterbilt": "Peter",
                "Kenworth": "Knwrth", "International": "Intl", "Volvo": "Volvo",
                "Mack": "Mack", "Isuzu": "Isuzu", "Hino": "Hino"}

BODY_TYPE_ORDER = [
    "Service Trucks", "Box Trucks", "Box Vans", "Flatbed Trucks",
    "Service Utility Vans", "Refrigerated", "Dump Trucks", "Cutaways",
    "Landscape Dumps", "Mechanic Body", "Stake Beds", "Contractor Trucks",
    "Enclosed Service", "Rollback", "Combo Trucks", "Bucket Trucks",
    "Chipper Trucks", "Dovetail Landscapes", "Hauler Body", "Hooklift",
    "Crane Trucks", "Wrecker", "Flatbed Dump", "Welder", "Roll-Off"
]

BODY_TYPE_ABBREV_HEADERS = [
    "Svc\nTruck", "Box\nTruck", "Box\nVan", "Flatbed", "Svc Util\nVan",
    "Refrig", "Dump", "Cutaway", "Lndscpe\nDump", "Mechanic",
    "Stake\nBed", "Contract", "Encl Svc", "Rollback", "Combo",
    "Bucket", "Chipper", "Dovetail", "Hauler", "Hooklift",
    "Crane", "Wrecker", "Flat\nDump", "Welder", "Roll-Off"
]

STATE_NAMES = {
    "AL": "Alabama", "AR": "Arkansas", "FL": "Florida", "GA": "Georgia",
    "KY": "Kentucky", "LA": "Louisiana", "MS": "Mississippi", "NC": "North Carolina",
    "OK": "Oklahoma", "SC": "South Carolina", "TN": "Tennessee", "TX": "Texas"
}


# ═══════════════════════════════════════════════════════════════════════════════
# LOGGING
# ═══════════════════════════════════════════════════════════════════════════════

LOG_FILE = None

def log(msg):
    ts = datetime.now().strftime('%H:%M:%S')
    line = f"[{ts}] {msg}"
    print(line, flush=True)
    if LOG_FILE:
        LOG_FILE.write(line + '\n')
        LOG_FILE.flush()


def log_section(title):
    log("")
    log("=" * 70)
    log(title)
    log("=" * 70)


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 1: COMVOY INVENTORY SCRAPER
# ═══════════════════════════════════════════════════════════════════════════════

def extract_vehicles_from_jsonld(soup, body_type_name, state):
    """Extract vehicle records from JSON-LD on a search results page."""
    vehicles = []

    for script in soup.find_all('script', type='application/ld+json'):
        try:
            data = json.loads(script.string)
        except (json.JSONDecodeError, TypeError):
            continue

        if not isinstance(data, dict):
            continue

        main_entity = data.get('mainEntity')
        if not isinstance(main_entity, dict):
            continue

        items = main_entity.get('itemListElement', [])
        if not items:
            continue

        vehicle_list = []
        for item in items:
            if not isinstance(item, dict):
                continue
            if item.get('type') == 'OfferCatalog' or item.get('@type') == 'OfferCatalog':
                vehicle_list.extend(item.get('itemListElement', []))
            elif item.get('type') == 'Vehicle' or item.get('@type') == 'Vehicle':
                vehicle_list.append(item)

        for v in vehicle_list:
            if not isinstance(v, dict):
                continue
            vtype = v.get('type', v.get('@type', ''))
            if vtype != 'Vehicle':
                continue

            vin = v.get('vehicleIdentificationNumber', '').strip()
            if not vin or len(vin) != 17:
                continue

            url = v.get('url', '')
            listing_id = ''
            url_match = re.search(r'-(\d{5,})$', url)
            if url_match:
                listing_id = url_match.group(1)

            name = v.get('name', '')
            condition = 'New' if name.lower().startswith('new') else 'Used'
            # Skip used vehicles — Comvoy only sells new
            if condition != 'New':
                continue
            year_match = re.search(r'(\d{4})', name)
            year = year_match.group(1) if year_match else ''

            dealer_name, city = '', ''
            dealer_state = state
            loc_match = re.search(r'/work-truck/([^/]+)-([a-z]{2})/', url)
            if loc_match:
                city = loc_match.group(1).replace('-', ' ').title()
                dealer_state = loc_match.group(2).upper()

            offers = v.get('offers', {})
            price = offers.get('price', '')

            brand_obj = v.get('brand', {})
            brand = brand_obj.get('name', '') if isinstance(brand_obj, dict) else ''
            model_obj = v.get('model', {})
            model = model_obj.get('name', '') if isinstance(model_obj, dict) else ''

            mfr = v.get('manufacturer', {})
            body_builder = mfr.get('name', '') if isinstance(mfr, dict) else ''

            vehicles.append({
                'vin': vin,
                'listing_id': listing_id,
                'condition': condition,
                'year': year,
                'brand': brand,
                'model': model,
                'body_type': body_type_name,
                'body_builder': body_builder,
                'price': price,
                'transmission': v.get('vehicleTransmission', ''),
                'fuel_type': v.get('fuelType', ''),
                'color': v.get('color', ''),
                'dealer_name': '',  # filled from HTML
                'city': city,
                'state': dealer_state,
                'listing_url': url,
                'image_url': v.get('image', ''),
            })

    return vehicles


def extract_dealers_from_html(soup):
    """
    Extract dealer entries from HTML -- one per vehicle card (1:1 with JSON-LD).
    Returns list of {'dealer_name', 'city', 'state'} in page order.
    """
    dealers = []
    headings = soup.find_all('div', class_=re.compile(r'category-heading'))

    for heading in headings:
        dealer_name = heading.get('title', '').strip() or heading.get_text(strip=True)
        if not dealer_name or dealer_name.lower() == 'key features':
            continue

        parent = heading.parent
        if not parent:
            continue

        location_div = parent.find('div', class_=re.compile(r'category-content'))
        if not location_div:
            gp = parent.parent
            if gp:
                location_div = gp.find('div', class_=re.compile(r'category-content'))

        location = ''
        if location_div:
            location = location_div.get('title', '').strip() or location_div.get_text(strip=True)

        if not re.match(r'^.+,\s*[A-Z]{2}$', location):
            continue

        parts = location.rsplit(',', 1)
        city, state = parts[0].strip(), parts[1].strip()
        dealers.append({'dealer_name': dealer_name, 'city': city, 'state': state})

    return dealers


def merge_dealers_into_vehicles(dealers, vehicles):
    """Positional zip: HTML dealer entries and JSON-LD vehicles 1:1."""
    for i, v in enumerate(vehicles):
        if i < len(dealers):
            v['dealer_name'] = dealers[i]['dealer_name']
            v['city'] = dealers[i]['city']
            v['state'] = dealers[i]['state']
    return vehicles


def get_total_results(soup):
    """Extract total result count from page."""
    text = soup.get_text()
    match = re.search(r'([\d,]+)\s+results?', text)
    if match:
        return int(match.group(1).replace(',', ''))
    return 0


def scrape_combo(session, brand_param, body_slug, body_name, state):
    """Scrape all pages for one brand+body+state combo."""
    base_url = f"https://www.comvoy.com/vehicles/{body_slug}?make={brand_param}&state={state}"

    try:
        resp = session.get(base_url, timeout=REQUEST_TIMEOUT)
        resp.raise_for_status()
    except requests.RequestException as e:
        return [], 0, str(e)

    soup = BeautifulSoup(resp.text, 'html.parser')
    total_results = get_total_results(soup)

    if total_results == 0:
        return [], 0, None

    total_pages = math.ceil(total_results / PER_PAGE)

    vehicles = extract_vehicles_from_jsonld(soup, body_name, state)
    dealers = extract_dealers_from_html(soup)
    merge_dealers_into_vehicles(dealers, vehicles)

    consecutive_empty = 0
    for page in range(2, total_pages + 1):
        time.sleep(DELAY)
        url = f"{base_url}&page={page}"
        try:
            resp = session.get(url, timeout=REQUEST_TIMEOUT)
            resp.raise_for_status()
        except requests.RequestException:
            continue

        soup = BeautifulSoup(resp.text, 'html.parser')
        page_vehicles = extract_vehicles_from_jsonld(soup, body_name, state)

        if not page_vehicles:
            consecutive_empty += 1
            if consecutive_empty >= MAX_EMPTY:
                break
        else:
            consecutive_empty = 0
            page_dealers = extract_dealers_from_html(soup)
            merge_dealers_into_vehicles(page_dealers, page_vehicles)
            vehicles.extend(page_vehicles)

    return vehicles, total_results, None


# ── Progress management ──────────────────────────────────────────────────────

def progress_path():
    return os.path.join(OUTPUT_DIR, 'scrape_progress.json')


def load_progress():
    p = progress_path()
    if os.path.exists(p):
        with open(p, 'r') as f:
            data = json.load(f)
            return set(data.get('completed', [])), data.get('vehicles_file', '')
    return set(), ''


def save_progress(completed, vehicles_file):
    with open(progress_path(), 'w') as f:
        json.dump({
            'completed': list(completed),
            'vehicles_file': vehicles_file,
            'last_updated': datetime.now().isoformat(),
        }, f)


def clear_progress():
    p = progress_path()
    if os.path.exists(p):
        os.remove(p)


def combo_key(brand_param, body_slug, state):
    return f"{brand_param}|{body_slug}|{state}"


def write_inventory_csv(path, vehicles_dict):
    """Write the master vehicle dict to CSV."""
    with open(path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=CSV_FIELDS, extrasaction='ignore')
        writer.writeheader()
        for v in vehicles_dict.values():
            writer.writerow(v)


def load_inventory_csv(path):
    """Load an inventory CSV into an OrderedDict keyed by VIN."""
    inventory = OrderedDict()
    if not os.path.exists(path):
        return inventory
    with open(path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            vin = row.get('vin', '').strip()
            if vin:
                inventory[vin] = row
    return inventory


def run_comvoy_scrape(date_str, inventory_path, resume=False):
    """Step 1: Scrape all Comvoy inventory. Returns OrderedDict of vehicles keyed by VIN."""
    log_section("STEP 1: SCRAPING COMVOY INVENTORY")

    total_combos = len(BRANDS) * len(BODY_TYPES) * len(STATES)
    log(f"Brands: {len(BRANDS)} | Bodies: {len(BODY_TYPES)} | States: {len(STATES)}")
    log(f"Total combos: {total_combos}")

    completed = set()
    if resume:
        completed, prev_file = load_progress()
        if prev_file and prev_file == inventory_path:
            log(f"Resuming: {len(completed)} combos already done")
        else:
            completed = set()

    # Master vehicle dict keyed by VIN (dedup across combos)
    all_vehicles = OrderedDict()

    # If resuming, load already-written vehicles
    if completed and os.path.exists(inventory_path):
        all_vehicles = load_inventory_csv(inventory_path)
        log(f"Loaded {len(all_vehicles)} vehicles from partial scrape")

    session = requests.Session()
    session.headers.update(HEADERS)

    combo_num = 0
    scraped_this_run = 0
    errors = 0
    start_time = time.time()

    for brand_param, brand_name in BRANDS.items():
        for body_slug, body_name in BODY_TYPES.items():
            for state in STATES:
                combo_num += 1
                key = combo_key(brand_param, body_slug, state)

                if key in completed:
                    continue

                scraped_this_run += 1
                remaining = total_combos - len(completed) - scraped_this_run
                elapsed = time.time() - start_time
                rate = scraped_this_run / elapsed * 60 if elapsed > 1 else 0
                eta = remaining / rate if rate > 0 else 0

                log(f"[{combo_num}/{total_combos}] {brand_name} | {body_name} | {state}  "
                    f"({len(all_vehicles)} vehicles, {rate:.1f}/min, ETA {eta:.0f}m)")

                vehicles, total_results, error = scrape_combo(
                    session, brand_param, body_slug, body_name, state
                )

                if error:
                    log(f"  ERROR: {error}")
                    errors += 1
                    time.sleep(DELAY)
                    continue

                new_count = 0
                for v in vehicles:
                    vin = v['vin']
                    # Skip excluded dealers (Penske, MHC, etc.)
                    dn = v.get('dealer_name', '').lower()
                    if any(pat in dn for pat in EXCLUDED_DEALER_PATTERNS):
                        continue
                    if vin not in all_vehicles:
                        v['scrape_date'] = date_str
                        all_vehicles[vin] = v
                        new_count += 1

                if vehicles:
                    log(f"  -> {len(vehicles)} on page, {new_count} new unique "
                        f"(Comvoy: {total_results} results)")

                completed.add(key)

                # Periodic save
                if scraped_this_run % PROGRESS_SAVE_EVERY == 0:
                    save_progress(completed, inventory_path)
                    write_inventory_csv(inventory_path, all_vehicles)

                time.sleep(DELAY)

    # Final write
    write_inventory_csv(inventory_path, all_vehicles)
    clear_progress()

    elapsed_total = time.time() - start_time
    log(f"Comvoy scrape complete: {len(all_vehicles)} unique vehicles")
    log(f"Combos scraped: {scraped_this_run} | Errors: {errors}")
    log(f"Time: {elapsed_total/60:.1f} minutes")

    return all_vehicles


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 2: SMYRNA / WORK TRUCK SOLUTIONS SCRAPER
# ═══════════════════════════════════════════════════════════════════════════════

def run_smyrna_scrape(date_str, smyrna_path):
    """Step 2: Scrape Smyrna WTS inventory. Returns set of VINs and list of vehicle dicts."""
    log_section("STEP 2: SCRAPING SMYRNA / WTS INVENTORY")

    all_vins = set()
    all_vehicles = []

    for page in range(1, SMYRNA_MAX_PAGES + 1):
        url = f'https://smyrnatruck.worktrucksolutions.com/vehicles?page={page}'
        try:
            resp = requests.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT)
            resp.raise_for_status()
        except requests.RequestException as e:
            log(f"  Page {page} error: {e}")
            break

        # Parse JSON-LD blocks (WTS uses semicolons after closing brace)
        ld_scripts = re.findall(
            r'<script[^>]*application/ld.json[^>]*>(.*?)</script>',
            resp.text, re.DOTALL
        )

        page_count = 0
        for s in ld_scripts:
            s = s.strip().rstrip(';')
            try:
                d = json.loads(s)
            except (json.JSONDecodeError, ValueError):
                continue

            if not isinstance(d, dict):
                continue

            tp = d.get('type', d.get('@type', ''))
            if tp != 'SearchResultsPage':
                continue

            me = d.get('mainEntity', {})
            if not isinstance(me, dict):
                continue

            for item in me.get('itemListElement', []):
                if not isinstance(item, dict):
                    continue
                for v in item.get('itemListElement', []):
                    if not isinstance(v, dict):
                        continue
                    vin = v.get('vehicleIdentificationNumber', '')
                    if vin and vin not in all_vins:
                        all_vins.add(vin)
                        brand = v.get('brand', {})
                        model = v.get('model', {})
                        mfr = v.get('manufacturer', {})
                        offers = v.get('offers', {})
                        all_vehicles.append({
                            'vin': vin,
                            'name': v.get('name', ''),
                            'price': offers.get('price', '') if isinstance(offers, dict) else '',
                            'brand': brand.get('name', '') if isinstance(brand, dict) else '',
                            'model': model.get('name', '') if isinstance(model, dict) else '',
                            'builder': mfr.get('name', '') if isinstance(mfr, dict) else '',
                        })
                        page_count += 1

        log(f"  Page {page}: {page_count} new vehicles (total: {len(all_vehicles)})")

        if page_count == 0:
            break
        time.sleep(SMYRNA_DELAY)

    # Write Smyrna CSV
    with open(smyrna_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=SMYRNA_CSV_FIELDS)
        writer.writeheader()
        writer.writerows(all_vehicles)

    log(f"Smyrna scrape complete: {len(all_vehicles)} unique vehicles")
    log(f"Output: {smyrna_path}")

    # Summary by builder
    builder_counts = Counter(v['builder'] for v in all_vehicles)
    for builder, count in builder_counts.most_common():
        log(f"  Builder: {builder or '(empty)'} = {count}")

    return all_vins, all_vehicles


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 3: DIFF AGAINST PREVIOUS SCRAPE
# ═══════════════════════════════════════════════════════════════════════════════

def find_previous_inventory(today_str):
    """Find the most recent inventory CSV before today."""
    pattern = re.compile(r'inventory_(\d{4}-\d{2}-\d{2})\.csv')
    candidates = []

    if not os.path.exists(OUTPUT_DIR):
        return None

    for fname in os.listdir(OUTPUT_DIR):
        m = pattern.match(fname)
        if m and m.group(1) != today_str:
            candidates.append((m.group(1), os.path.join(OUTPUT_DIR, fname)))

    if not candidates:
        return None

    candidates.sort(reverse=True)
    return candidates[0][1]


def run_diff(date_str, current_path):
    """Step 3: Diff current vs previous inventory."""
    log_section("STEP 3: DIFF AGAINST PREVIOUS SCRAPE")

    prev_path = find_previous_inventory(date_str)
    if not prev_path:
        log("No previous inventory found -- skipping diff")
        return None

    log(f"Previous scrape: {os.path.basename(prev_path)}")

    current = load_inventory_csv(current_path)
    previous = load_inventory_csv(prev_path)

    current_vins = set(current.keys())
    previous_vins = set(previous.keys())

    new_vins = current_vins - previous_vins
    sold_vins = previous_vins - current_vins
    common_vins = current_vins & previous_vins

    # Price changes
    price_changes = []
    for vin in common_vins:
        old_price = previous[vin].get('price', '')
        new_price = current[vin].get('price', '')
        if old_price and new_price and old_price != new_price:
            row = dict(current[vin])
            row['old_price'] = old_price
            row['new_price'] = new_price
            try:
                row['price_diff'] = str(int(new_price) - int(old_price))
            except ValueError:
                row['price_diff'] = ''
            price_changes.append(row)

    # Write new vehicles
    new_path = os.path.join(OUTPUT_DIR, f'new_vehicles_{date_str}.csv')
    with open(new_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=CSV_FIELDS)
        writer.writeheader()
        for vin in sorted(new_vins):
            writer.writerow(current[vin])

    # Write sold vehicles
    sold_path = os.path.join(OUTPUT_DIR, f'sold_vehicles_{date_str}.csv')
    with open(sold_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=CSV_FIELDS)
        writer.writeheader()
        for vin in sorted(sold_vins):
            writer.writerow(previous[vin])

    # Write price changes
    pc_fields = CSV_FIELDS + ['old_price', 'new_price', 'price_diff']
    pc_path = os.path.join(OUTPUT_DIR, f'price_changes_{date_str}.csv')
    with open(pc_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=pc_fields)
        writer.writeheader()
        for row in price_changes:
            writer.writerow(row)

    diff_stats = {
        'previous_file': os.path.basename(prev_path),
        'previous_count': len(previous),
        'current_count': len(current),
        'new': len(new_vins),
        'sold': len(sold_vins),
        'price_changes': len(price_changes),
        'retained': len(common_vins) - len(price_changes),
    }

    log(f"Previous vehicles:  {diff_stats['previous_count']:,}")
    log(f"Current vehicles:   {diff_stats['current_count']:,}")
    log(f"New vehicles:       {diff_stats['new']:,}")
    log(f"Sold vehicles:      {diff_stats['sold']:,}")
    log(f"Price changes:      {diff_stats['price_changes']:,}")
    log(f"Retained (same):    {diff_stats['retained']:,}")

    return diff_stats


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 4: BUILD EXCEL REPORT
# ═══════════════════════════════════════════════════════════════════════════════

def calibri(size=11, bold=False, italic=False, color="00000000"):
    return Font(name="Calibri", size=size, bold=bold, italic=italic, color=color)


def set_header_row(ws, row, columns, col_start=1):
    for i, text in enumerate(columns):
        cell = ws.cell(row=row, column=col_start + i, value=text)
        cell.font = calibri(11, bold=True, color=WHITE)
        cell.fill = navy_fill
        cell.alignment = center_wrap
        cell.border = thin_border


def set_header_row_small(ws, row, columns, col_start=1, font_size=9):
    for i, text in enumerate(columns):
        cell = ws.cell(row=row, column=col_start + i, value=text)
        cell.font = calibri(font_size, bold=True, color=WHITE)
        cell.fill = navy_fill
        cell.alignment = center_wrap
        cell.border = thin_border


def apply_data_cell(cell, font_size=11, bold=False, color="00000000",
                    align=None, fill=None, num_fmt=None):
    cell.font = calibri(font_size, bold=bold, color=color)
    cell.border = thin_border
    if align:
        cell.alignment = align
    if fill:
        cell.fill = fill
    if num_fmt:
        cell.number_format = num_fmt


def run_report(date_str, inventory_path, smyrna_vins, smyrna_vehicles, diff_stats, old_report_path):
    """Step 4: Build the full 9-sheet Excel report."""
    log_section("STEP 4: BUILDING EXCEL REPORT")

    report_path = os.path.join(OUTPUT_DIR, f'Comvoy_Multi_Brand_Report_{date_str}.xlsx')
    report_date = datetime.strptime(date_str, '%Y-%m-%d')
    date_display = report_date.strftime('%B %d, %Y')
    month_display = report_date.strftime('%B %Y')

    # Load inventory
    log("Loading inventory CSV...")
    with open(inventory_path, encoding='utf-8') as f:
        reader = csv.DictReader(f)
        vehicles = list(reader)
    log(f"  Loaded {len(vehicles):,} vehicles")

    for v in vehicles:
        try:
            v['price_num'] = int(float(v['price'])) if v['price'] else 0
        except (ValueError, TypeError):
            v['price_num'] = 0

    # ── Build aggregations ────────────────────────────────────────────────
    # Dealer grouping: ALWAYS by (dealer_name, city, state)
    dealer_data = defaultdict(lambda: {"vehicles": [], "brands": set(), "body_types": set(),
                                        "smyrna_count": 0, "smyrna_vin_count": 0})

    for v in vehicles:
        key = (v['dealer_name'], v['city'], v['state'])
        d = dealer_data[key]
        d['vehicles'].append(v)
        d['brands'].add(v['brand'])
        d['body_types'].add(v['body_type'])
        # body_builder-based Smyrna count (legacy)
        if v['body_builder'] == 'Smyrna Truck':
            d['smyrna_count'] += 1
        # VIN-based Smyrna count (source of truth from WTS scrape)
        if v['vin'] in smyrna_vins:
            d['smyrna_vin_count'] += 1

    total_vehicles = len(vehicles)
    total_dealers = len(dealer_data)
    total_brands = len(set(v['brand'] for v in vehicles))
    total_body_types = len(set(v['body_type'] for v in vehicles))

    # Use VIN-based Smyrna counts as source of truth
    total_smyrna = sum(d['smyrna_vin_count'] for d in dealer_data.values())
    smyrna_dealers = sum(1 for d in dealer_data.values() if d['smyrna_vin_count'] > 0)

    dealers_sorted = sorted(dealer_data.items(), key=lambda x: -len(x[1]['vehicles']))

    # Brand aggregation
    brand_data = defaultdict(lambda: {"dealers": set(), "vehicles": [], "body_types": Counter()})
    for v in vehicles:
        bd = brand_data[v['brand']]
        bd['dealers'].add((v['dealer_name'], v['city'], v['state']))
        bd['vehicles'].append(v)
        bd['body_types'][v['body_type']] += 1
    brands_sorted = sorted(brand_data.items(), key=lambda x: -len(x[1]['vehicles']))

    # Body type aggregation
    bt_data = defaultdict(lambda: {"dealers": set(), "vehicles": [], "smyrna": 0, "smyrna_vin": 0})
    for v in vehicles:
        btd = bt_data[v['body_type']]
        btd['dealers'].add((v['dealer_name'], v['city'], v['state']))
        btd['vehicles'].append(v)
        if v['body_builder'] == 'Smyrna Truck':
            btd['smyrna'] += 1
        if v['vin'] in smyrna_vins:
            btd['smyrna_vin'] += 1
    bts_sorted = sorted(bt_data.items(), key=lambda x: -len(x[1]['vehicles']))

    # State aggregation
    state_data = defaultdict(lambda: {"dealers": set(), "vehicles": 0, "brands": set()})
    for v in vehicles:
        sd = state_data[v['state']]
        sd['dealers'].add((v['dealer_name'], v['city'], v['state']))
        sd['vehicles'] += 1
        sd['brands'].add(v['brand'])
    states_sorted = sorted(state_data.items(), key=lambda x: -x[1]['vehicles'])

    # Load old report for dealer metrics comparison (if it exists)
    old_dealer_counts = {}
    if old_report_path and os.path.exists(old_report_path):
        log(f"Loading old report for comparison: {os.path.basename(old_report_path)}")
        try:
            old_wb = load_workbook(old_report_path, read_only=True, data_only=True)
            if "All Dealers" in old_wb.sheetnames:
                old_ws = old_wb["All Dealers"]
                for row in old_ws.iter_rows(min_row=5, values_only=True):
                    if row[1] and row[4]:
                        name = str(row[1]).strip()
                        city = str(row[2] or "").strip()
                        state = str(row[3] or "").strip()
                        count = int(row[4]) if row[4] else 0
                        old_dealer_counts[(name, city, state)] = count
            old_wb.close()
            log(f"  Loaded {len(old_dealer_counts)} old dealers")
        except Exception as e:
            log(f"  Warning: Could not load old report: {e}")
    else:
        # Try to find a previous report automatically
        prev_inventory = find_previous_inventory(date_str)
        if prev_inventory:
            prev_date = re.search(r'inventory_(\d{4}-\d{2}-\d{2})', prev_inventory)
            if prev_date:
                prev_report = os.path.join(OUTPUT_DIR,
                    f'Comvoy_Multi_Brand_Report_{prev_date.group(1)}.xlsx')
                if os.path.exists(prev_report):
                    log(f"Loading previous report for comparison: {os.path.basename(prev_report)}")
                    try:
                        old_wb = load_workbook(prev_report, read_only=True, data_only=True)
                        if "All Dealers" in old_wb.sheetnames:
                            old_ws = old_wb["All Dealers"]
                            for row in old_ws.iter_rows(min_row=5, values_only=True):
                                if row[1] and row[4]:
                                    name = str(row[1]).strip()
                                    city = str(row[2] or "").strip()
                                    state = str(row[3] or "").strip()
                                    count = int(row[4]) if row[4] else 0
                                    old_dealer_counts[(name, city, state)] = count
                        old_wb.close()
                        log(f"  Loaded {len(old_dealer_counts)} old dealers")
                    except Exception as e:
                        log(f"  Warning: Could not load previous report: {e}")

    # ── Create Workbook ───────────────────────────────────────────────────
    wb = Workbook()

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 1: Executive Summary
    # ══════════════════════════════════════════════════════════════════════
    log("  Sheet 1: Executive Summary...")
    ws = wb.active
    ws.title = "Executive Summary"

    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = "Comvoy Territory Intelligence Report"
    cell.font = calibri(22, bold=True, color=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 45

    ws.merge_cells("A2:H2")
    cell = ws["A2"]
    cell.value = f"Generated {date_display} | 12 States | Comvoy.com Inventory Analysis"
    cell.font = calibri(12, italic=True, color=GRAY)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 25

    ws.row_dimensions[3].height = 6
    navy_bottom = Border(bottom=Side(style="medium", color=NAVY))
    for col in range(1, 9):
        ws.cell(row=3, column=col).border = navy_bottom

    ws.row_dimensions[4].height = 8

    ws.row_dimensions[5].height = 50
    ws.row_dimensions[6].height = 22

    kpis = [
        (1, total_dealers, "Territory Dealers"),
        (3, total_vehicles, "Total Vehicles"),
        (5, total_brands, "Chassis Brands"),
        (7, total_body_types, "Body Types"),
    ]
    for col, value, label in kpis:
        cell = ws.cell(row=5, column=col, value=value)
        cell.font = calibri(28, bold=True, color=NAVY)
        cell.alignment = center
        cell = ws.cell(row=6, column=col, value=label)
        cell.font = calibri(10, color=GRAY)
        cell.alignment = center

    ws.row_dimensions[7].height = 6
    thin_bottom = Border(bottom=Side(style="thin", color=LIGHT_BLUE))
    for col in range(1, 9):
        ws.cell(row=7, column=col).border = thin_bottom

    ws.row_dimensions[8].height = 8

    ws.merge_cells("A9:H9")
    smyrna_pct = total_smyrna / total_vehicles * 100 if total_vehicles else 0
    cell = ws["A9"]
    cell.value = (f"Smyrna Truck / Fouts Bros Products: {total_smyrna} vehicles "
                  f"at {smyrna_dealers} dealers ({smyrna_pct:.2f}% of territory)")
    cell.font = calibri(12, bold=True, color=BROWN)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[9].height = 28

    ws.row_dimensions[10].height = 8

    ws.row_dimensions[11].height = 28
    ws.merge_cells("A11:D11")
    cell = ws["A11"]
    cell.value = "Top Brands by Volume (Territory)"
    cell.font = calibri(13, bold=True, color=NAVY)

    ws.merge_cells("F11:H11")
    cell = ws["F11"]
    cell.value = "Top Dealers by Volume"
    cell.font = calibri(13, bold=True, color=NAVY)

    set_header_row(ws, 12, ["Brand", "Vehicles", "Dealers", "% of Total"], col_start=1)
    set_header_row(ws, 12, ["Dealer", "Vehicles", "Smyrna"], col_start=6)

    for i, (brand, bd) in enumerate(brands_sorted[:7]):
        row = 13 + i
        ws.row_dimensions[row].height = 22
        fill = zebra_fill if i % 2 == 0 else None
        pct_str = f"{len(bd['vehicles']) / total_vehicles * 100:.1f}%"
        vals = [brand, len(bd['vehicles']), len(bd['dealers']), pct_str]
        for j, val in enumerate(vals):
            cell = ws.cell(row=row, column=1 + j, value=val)
            cell.font = calibri(11)
            if fill:
                cell.fill = fill

    for i, ((name, city, state), dd) in enumerate(dealers_sorted[:7]):
        row = 13 + i
        fill = zebra_fill if i % 2 == 0 else None
        smyrna_val = dd['smyrna_vin_count'] if dd['smyrna_vin_count'] > 0 else None

        cell = ws.cell(row=row, column=6, value=name)
        cell.font = calibri(11)
        if fill:
            cell.fill = fill
        cell = ws.cell(row=row, column=7, value=len(dd['vehicles']))
        cell.font = calibri(11)
        if fill:
            cell.fill = fill
        if smyrna_val:
            cell = ws.cell(row=row, column=8, value=smyrna_val)
            cell.font = calibri(11)
            if fill:
                cell.fill = fill

    ws.row_dimensions[20].height = 8

    ws.merge_cells("A21:E21")
    cell = ws["A21"]
    cell.value = "12-State Territory Breakdown"
    cell.font = calibri(13, bold=True, color=NAVY)
    ws.row_dimensions[21].height = 28

    set_header_row(ws, 22, ["State", "Dealers", "Vehicles", "Brands", "% of Territory"], col_start=1)

    for i, (st, sd) in enumerate(states_sorted):
        row = 23 + i
        ws.row_dimensions[row].height = 22
        fill = zebra_fill if i % 2 == 0 else None
        pct_str = f"{sd['vehicles'] / total_vehicles * 100:.1f}%"
        state_name = STATE_NAMES.get(st, st)
        vals = [state_name, len(sd['dealers']), sd['vehicles'], len(sd['brands']), pct_str]
        for j, val in enumerate(vals):
            cell = ws.cell(row=row, column=1 + j, value=val)
            cell.font = calibri(11)
            if fill:
                cell.fill = fill

    total_row = 23 + len(states_sorted)
    ws.row_dimensions[total_row].height = 28
    total_vals = ["TERRITORY TOTAL", total_dealers, total_vehicles, total_brands, "100%"]
    for j, val in enumerate(total_vals):
        cell = ws.cell(row=total_row, column=1 + j, value=val)
        cell.font = calibri(12, bold=True, color=WHITE)
        cell.fill = navy_fill

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 38
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 10

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 2: Brand Summary
    # ══════════════════════════════════════════════════════════════════════
    log("  Sheet 2: Brand Summary...")
    ws = wb.create_sheet("Brand Summary")

    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = f"Inventory by Chassis Brand \u2014 12-State Territory"
    cell.font = calibri(16, bold=True, color=NAVY)
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    cell = ws["A2"]
    cell.value = f"{total_brands} brands across {total_dealers} dealers | {total_vehicles:,} total vehicles"
    cell.font = calibri(11, italic=True, color=GRAY)
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 8

    set_header_row(ws, 4, ["Brand", "Dealers", "Total Vehicles", "Avg per Dealer",
                            "% of Territory", "Top Body Type", "Top Body Vehicles"])
    ws.row_dimensions[4].height = 28

    for i, (brand, bd) in enumerate(brands_sorted):
        row = 5 + i
        ws.row_dimensions[row].height = 22
        fill = zebra_fill if i % 2 == 0 else None
        n_vehicles = len(bd['vehicles'])
        n_dealers = len(bd['dealers'])
        avg = round(n_vehicles / n_dealers, 1) if n_dealers else 0
        pct = n_vehicles / total_vehicles if total_vehicles else 0
        top_bt = bd['body_types'].most_common(1)[0] if bd['body_types'] else ("", 0)

        vals = [brand, n_dealers, n_vehicles, avg, pct, top_bt[0], top_bt[1]]
        fmts = [None, "#,##0", "#,##0", "#,##0.0", "0.0%", None, "#,##0"]
        aligns = [left_center, center, center, center, center, left_center, center]

        for j, (val, fmt, al) in enumerate(zip(vals, fmts, aligns)):
            cell = ws.cell(row=row, column=1 + j, value=val)
            apply_data_cell(cell, align=al, fill=fill, num_fmt=fmt)

    total_row = 5 + len(brands_sorted)
    ws.row_dimensions[total_row].height = 28
    cell = ws.cell(row=total_row, column=1, value="TOTAL")
    cell.font = calibri(11, bold=True)
    cell.border = thin_border
    cell = ws.cell(row=total_row, column=2, value=total_dealers)
    cell.font = calibri(11, bold=True)
    cell.border = thin_border
    cell.number_format = "#,##0"
    cell = ws.cell(row=total_row, column=3, value=total_vehicles)
    cell.font = calibri(11, bold=True)
    cell.border = thin_border
    cell.number_format = "#,##0"
    for j in range(4, 8):
        ws.cell(row=total_row, column=j).font = calibri(11, bold=True)
        ws.cell(row=total_row, column=j).border = thin_border

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 16

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 3: Body Type Summary
    # ══════════════════════════════════════════════════════════════════════
    log("  Sheet 3: Body Type Summary...")
    ws = wb.create_sheet("Body Type Summary")

    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = f"Inventory by Body Type \u2014 12-State Territory"
    cell.font = calibri(16, bold=True, color=NAVY)
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    cell = ws["A2"]
    cell.value = f"{total_body_types} active body types | All {total_brands} brands combined"
    cell.font = calibri(11, italic=True, color=GRAY)
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 8

    set_header_row(ws, 4, ["Body Type", "Dealers", "Total Vehicles", "Avg per Dealer",
                            "% of Territory", "Smyrna Vehicles"])
    ws.row_dimensions[4].height = 28

    for i, (bt, btd) in enumerate(bts_sorted):
        row = 5 + i
        ws.row_dimensions[row].height = 22
        fill = zebra_fill if i % 2 == 0 else None
        n_vehicles = len(btd['vehicles'])
        n_dealers = len(btd['dealers'])
        avg = round(n_vehicles / n_dealers, 1) if n_dealers else 0
        pct = n_vehicles / total_vehicles if total_vehicles else 0
        # Use VIN-based Smyrna count
        smyrna_v = btd['smyrna_vin'] if btd['smyrna_vin'] > 0 else None

        vals = [bt, n_dealers, n_vehicles, avg, pct, smyrna_v]
        fmts = [None, "#,##0", "#,##0", "#,##0.0", "0.0%", "#,##0"]
        aligns = [left_center, center, center, center, center, center]

        for j, (val, fmt, al) in enumerate(zip(vals, fmts, aligns)):
            cell = ws.cell(row=row, column=1 + j, value=val)
            apply_data_cell(cell, align=al, fill=fill, num_fmt=fmt)

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 18

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 4: All Dealers
    # ══════════════════════════════════════════════════════════════════════
    log("  Sheet 4: All Dealers...")
    ws = wb.create_sheet("All Dealers")

    ws.merge_cells("A1:K1")
    cell = ws["A1"]
    cell.value = f"Complete Dealer Listing \u2014 {total_dealers} Dealers Across 12 States"
    cell.font = calibri(16, bold=True, color=NAVY)
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:K2")
    cell = ws["A2"]
    cell.value = ("Ranked by total vehicles | Gold rows = Top 100 | "
                  "Smyrna % = Smyrna/Fouts products as share of dealer inventory")
    cell.font = calibri(11, italic=True, color=GRAY)
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 8

    set_header_row(ws, 4, ["Rank", "Dealer Name", "City", "ST", "Total\nVehicles", "Brands",
                            "Body\nTypes", "Top Brand", "Smyrna\nUnits", "Smyrna\n%", "Top Body Types"])
    ws.row_dimensions[4].height = 28

    for i, ((name, city, state), dd) in enumerate(dealers_sorted):
        row = 5 + i
        ws.row_dimensions[row].height = 22
        rank = i + 1
        is_top100 = rank <= 100

        n_vehicles = len(dd['vehicles'])
        n_brands = len(dd['brands'])
        n_body_types = len(dd['body_types'])

        brand_counter = Counter(v['brand'] for v in dd['vehicles'])
        top_brand_name, top_brand_count = brand_counter.most_common(1)[0]
        top_brand_str = f"{top_brand_name} ({top_brand_count:,})"

        # Use VIN-based Smyrna count
        smyrna_units = dd['smyrna_vin_count'] if dd['smyrna_vin_count'] > 0 else None
        smyrna_pct_val = dd['smyrna_vin_count'] / n_vehicles if dd['smyrna_vin_count'] > 0 else None

        bt_counter = Counter(v['body_type'] for v in dd['vehicles'])
        top_bts = ", ".join(k for k, _ in bt_counter.most_common(3))

        vals = [rank, name, city, state, n_vehicles, n_brands, n_body_types,
                top_brand_str, smyrna_units, smyrna_pct_val, top_bts]
        fmts = [None, None, None, None, "#,##0", None, None, None, "#,##0", "0.0%", None]
        aligns_row = [center, left_center, left_center, center, center, center, center,
                      left_center, center, center, left_wrap]

        for j, (val, fmt, al) in enumerate(zip(vals, fmts, aligns_row)):
            cell = ws.cell(row=row, column=1 + j, value=val)
            cell.font = calibri(11, bold=(j == 0), color=NAVY if j == 0 else "00000000")
            cell.border = thin_border
            cell.alignment = al
            if fmt and val is not None:
                cell.number_format = fmt
            if is_top100:
                cell.fill = gold_fill

    for col_letter, width in zip("ABCDEFGHIJK", [7, 42, 18, 5, 12, 9, 9, 22, 10, 10, 48]):
        ws.column_dimensions[col_letter].width = width

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 5: Dealer Brand Matrix
    # ══════════════════════════════════════════════════════════════════════
    log("  Sheet 5: Dealer Brand Matrix...")
    ws = wb.create_sheet("Dealer Brand Matrix")

    ws.merge_cells("A1:R1")
    cell = ws["A1"]
    cell.value = f"All {total_dealers} Dealers \u2014 Chassis Brand Matrix"
    cell.font = calibri(16, bold=True, color=NAVY)
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:R2")
    cell = ws["A2"]
    cell.value = f"Vehicles by chassis make | Green = has inventory | Gold rows = Top 100 | {month_display}"
    cell.font = calibri(11, italic=True, color=GRAY)
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 8

    headers = ["Dealer Name", "City", "ST", "Brands"] + \
              [BRAND_ABBREV.get(b, b) for b in BRAND_ORDER] + ["Total"]
    ws.row_dimensions[4].height = 50

    for i, text in enumerate(headers):
        cell = ws.cell(row=4, column=1 + i, value=text)
        cell.fill = navy_fill
        cell.alignment = center_wrap
        cell.border = thin_border
        if i < 4 or i == len(headers) - 1:
            cell.font = calibri(11, bold=True, color=WHITE)
        else:
            cell.font = calibri(9, bold=True, color=WHITE)

    for i, ((name, city, state), dd) in enumerate(dealers_sorted):
        row = 5 + i
        ws.row_dimensions[row].height = 18
        rank = i + 1
        is_top100 = rank <= 100

        brand_counts = Counter(v['brand'] for v in dd['vehicles'])
        n_brands = len(brand_counts)
        n_total = len(dd['vehicles'])

        for j, val in enumerate([name, city, state]):
            cell = ws.cell(row=row, column=1 + j, value=val)
            cell.font = calibri(9)
            cell.border = thin_border
            cell.alignment = left_center if j < 2 else center
            if is_top100:
                cell.fill = gold_fill

        cell = ws.cell(row=row, column=4, value=n_brands)
        cell.font = calibri(9)
        cell.border = thin_border
        cell.alignment = center
        cell.fill = blue_fill if not is_top100 else gold_fill

        for j, brand in enumerate(BRAND_ORDER):
            col = 5 + j
            count = brand_counts.get(brand, 0)
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = center
            if count > 0:
                cell.value = count
                cell.font = calibri(9, bold=True, color=NAVY)
                if not is_top100:
                    cell.fill = green_fill
                else:
                    cell.fill = gold_fill
            else:
                cell.font = calibri(9)
                if is_top100:
                    cell.fill = gold_fill

        cell = ws.cell(row=row, column=5 + len(BRAND_ORDER), value=n_total)
        cell.font = calibri(9, bold=True)
        cell.border = thin_border
        cell.alignment = center
        if is_top100:
            cell.fill = gold_fill
        else:
            cell.fill = blue_fill

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 5
    ws.column_dimensions["D"].width = 8
    for j in range(len(BRAND_ORDER)):
        ws.column_dimensions[get_column_letter(5 + j)].width = 8
    ws.column_dimensions[get_column_letter(5 + len(BRAND_ORDER))].width = 9

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 6: Dealer Body Type Matrix
    # ══════════════════════════════════════════════════════════════════════
    log("  Sheet 6: Dealer Body Type Matrix...")
    ws = wb.create_sheet("Dealer Body Type Matrix")

    n_bt_cols = len(BODY_TYPE_ORDER)
    last_col = 5 + n_bt_cols
    last_col_letter = get_column_letter(last_col)

    ws.merge_cells(f"A1:{last_col_letter}1")
    cell = ws["A1"]
    cell.value = f"All {total_dealers} Dealers \u2014 Body Type Matrix"
    cell.font = calibri(16, bold=True, color=NAVY)
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f"A2:{last_col_letter}2")
    cell = ws["A2"]
    cell.value = f"Vehicles by body type | Green = has inventory | Gold rows = Top 100 | {month_display}"
    cell.font = calibri(11, italic=True, color=GRAY)
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 8

    ws.row_dimensions[4].height = 50
    base_headers = ["Dealer Name", "City", "ST", "Types"]
    for i, text in enumerate(base_headers):
        cell = ws.cell(row=4, column=1 + i, value=text)
        cell.font = calibri(11, bold=True, color=WHITE)
        cell.fill = navy_fill
        cell.alignment = center_wrap
        cell.border = thin_border

    for i, text in enumerate(BODY_TYPE_ABBREV_HEADERS):
        cell = ws.cell(row=4, column=5 + i, value=text)
        cell.font = calibri(8, bold=True, color=WHITE)
        cell.fill = navy_fill
        cell.alignment = center_wrap
        cell.border = thin_border

    cell = ws.cell(row=4, column=last_col, value="Total")
    cell.font = calibri(11, bold=True, color=WHITE)
    cell.fill = navy_fill
    cell.alignment = center_wrap
    cell.border = thin_border

    for i, ((name, city, state), dd) in enumerate(dealers_sorted):
        row = 5 + i
        ws.row_dimensions[row].height = 18
        rank = i + 1
        is_top100 = rank <= 100

        bt_counts = Counter(v['body_type'] for v in dd['vehicles'])
        n_types = len(bt_counts)
        n_total = len(dd['vehicles'])

        for j, val in enumerate([name, city, state]):
            cell = ws.cell(row=row, column=1 + j, value=val)
            cell.font = calibri(9)
            cell.border = thin_border
            cell.alignment = left_center if j < 2 else center
            if is_top100:
                cell.fill = gold_fill

        cell = ws.cell(row=row, column=4, value=n_types)
        cell.font = calibri(9)
        cell.border = thin_border
        cell.alignment = center
        cell.fill = blue_fill if not is_top100 else gold_fill

        for j, bt in enumerate(BODY_TYPE_ORDER):
            col = 5 + j
            count = bt_counts.get(bt, 0)
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = center
            if count > 0:
                cell.value = count
                cell.font = calibri(9, bold=True, color=NAVY)
                if not is_top100:
                    cell.fill = green_fill
                else:
                    cell.fill = gold_fill
            else:
                cell.font = calibri(9)
                if is_top100:
                    cell.fill = gold_fill

        cell = ws.cell(row=row, column=last_col, value=n_total)
        cell.font = calibri(9, bold=True)
        cell.border = thin_border
        cell.alignment = center
        if is_top100:
            cell.fill = gold_fill
        else:
            cell.fill = blue_fill

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 5
    ws.column_dimensions["D"].width = 8
    for j in range(n_bt_cols):
        ws.column_dimensions[get_column_letter(5 + j)].width = 8
    ws.column_dimensions[last_col_letter].width = 9

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 7: Smyrna Product Analysis (VIN-based source of truth)
    # ══════════════════════════════════════════════════════════════════════
    log("  Sheet 7: Smyrna Product Analysis (VIN-based)...")
    ws = wb.create_sheet("Smyrna Product Analysis")

    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = "Smyrna Truck / Fouts Bros \u2014 Product Placement Analysis"
    cell.font = calibri(16, bold=True, color=BROWN)
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    cell = ws["A2"]
    cell.value = (f"{total_smyrna} Smyrna/Fouts vehicles across {smyrna_dealers} dealers | "
                  f"VIN-matched from smyrnatruck.worktrucksolutions.com ({len(smyrna_vins)} WTS VINs)")
    cell.font = calibri(11, italic=True, color=GRAY)
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 8

    set_header_row(ws, 4, ["Rank", "Dealer Name", "City", "ST", "Smyrna\nUnits",
                            "Dealer\nTotal", "Smyrna %", "Top Smyrna Body Types"])
    ws.row_dimensions[4].height = 28

    # Get Smyrna dealers sorted by VIN-based smyrna count desc
    smyrna_dealer_list = [(key, dd) for key, dd in dealers_sorted if dd['smyrna_vin_count'] > 0]
    smyrna_dealer_list.sort(key=lambda x: -x[1]['smyrna_vin_count'])

    for i, ((name, city, state), dd) in enumerate(smyrna_dealer_list):
        row = 5 + i
        ws.row_dimensions[row].height = 22
        fill = peach_fill if i % 2 == 0 else None
        rank = i + 1

        # Smyrna vehicles identified by VIN match
        smyrna_vehicles_here = [v for v in dd['vehicles'] if v['vin'] in smyrna_vins]
        smyrna_bt = Counter(v['body_type'] for v in smyrna_vehicles_here)
        top_smyrna_bts = ", ".join(f"{k} ({v})" for k, v in smyrna_bt.most_common(3))

        n_total = len(dd['vehicles'])
        smyrna_pct_val = dd['smyrna_vin_count'] / n_total if n_total else 0

        vals = [rank, name, city, state, dd['smyrna_vin_count'],
                n_total, smyrna_pct_val, top_smyrna_bts]
        fmts = [None, None, None, None, "#,##0", "#,##0", "0.0%", None]
        aligns_row = [center, left_center, left_center, center, center, center, center, left_wrap]

        for j, (val, fmt, al) in enumerate(zip(vals, fmts, aligns_row)):
            cell = ws.cell(row=row, column=1 + j, value=val)
            cell.font = calibri(11, bold=(j == 0), color=NAVY if j == 0 else "00000000")
            cell.border = thin_border
            cell.alignment = al
            if fmt and val is not None:
                cell.number_format = fmt
            if fill:
                cell.fill = fill

    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 5
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 48

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 8: Dealer Metrics (comparison with previous month)
    # ══════════════════════════════════════════════════════════════════════
    log("  Sheet 8: Dealer Metrics...")
    ws = wb.create_sheet("Dealer Metrics")

    ws.merge_cells("A1:I1")
    cell = ws["A1"]
    cell.value = "Dealer Inventory Changes"
    cell.font = calibri(16, bold=True, color=NAVY)
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:I2")
    cell = ws["A2"]
    if old_dealer_counts:
        cell.value = f"Comparison vs Previous Month | Matched by Dealer + City + State | {month_display}"
    else:
        cell.value = f"No previous report available for comparison | {month_display}"
    cell.font = calibri(11, italic=True, color=GRAY)
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 8

    if old_dealer_counts:
        # Build comparison
        new_dealer_keys = {}
        for key, dd in dealer_data.items():
            new_dealer_keys[key] = len(dd['vehicles'])

        all_keys = set(old_dealer_counts.keys()) | set(new_dealer_keys.keys())
        comparison = []
        surges = declines = new_count = lost_count = 0

        for key in all_keys:
            name, city, state = key
            old_val = old_dealer_counts.get(key, 0)
            new_val = new_dealer_keys.get(key, 0)
            change = new_val - old_val
            abs_change = abs(change)
            pct_change = change / old_val if old_val > 0 else (None if new_val > 0 else 0)

            status = ""
            if key not in old_dealer_counts:
                status = "NEW"
                new_count += 1
            elif key not in new_dealer_keys:
                status = "LOST"
                lost_count += 1
            elif old_val > 5 and pct_change is not None and pct_change > 0.5:
                status = "SURGE"
                surges += 1
            elif old_val > 5 and pct_change is not None and pct_change < -0.5:
                status = "DECLINE"
                declines += 1

            comparison.append((name, city, state, old_val, new_val, change, pct_change, status, abs_change))

        comparison.sort(key=lambda x: -x[8])

        # Summary KPIs
        ws.row_dimensions[4].height = 40
        ws.row_dimensions[5].height = 18
        kpi_data = [
            (1, surges, "SURGE", green_fill),
            (3, declines, "DECLINE", light_red_fill),
            (5, new_count, "NEW", light_yellow_fill),
            (7, lost_count, "LOST", light_gray_fill),
        ]
        for col, val, label, fill in kpi_data:
            cell = ws.cell(row=4, column=col, value=val)
            cell.font = calibri(28, bold=True, color=NAVY)
            cell.alignment = center
            cell.fill = fill
            cell = ws.cell(row=5, column=col, value=label)
            cell.font = calibri(10, color=GRAY)
            cell.alignment = center
            cell.fill = fill

        ws.row_dimensions[6].height = 8

        # Headers row 7
        set_header_row(ws, 7, ["Rank", "Dealer Name", "City", "ST", "Old Count",
                                "New Count", "Change", "% Change", "Status"])
        ws.row_dimensions[7].height = 28

        GREEN_FONT = Font(name="Calibri", size=11, color="006100")
        RED_FONT = Font(name="Calibri", size=11, color="CC0000")

        for i, (name, city, state, old_val, new_val, change, pct_change, status, _) in enumerate(comparison):
            r = 8 + i
            ws.row_dimensions[r].height = 22

            ws.cell(row=r, column=1, value=i + 1).font = calibri(11, bold=True, color=NAVY)
            ws.cell(row=r, column=2, value=name)
            ws.cell(row=r, column=3, value=city)
            ws.cell(row=r, column=4, value=state)

            old_cell = ws.cell(row=r, column=5, value=old_val if old_val else None)
            old_cell.number_format = "#,##0"

            new_cell = ws.cell(row=r, column=6, value=new_val if new_val else None)
            new_cell.number_format = "#,##0"

            chg_cell = ws.cell(row=r, column=7, value=change)
            if change > 0:
                chg_cell.font = GREEN_FONT
            elif change < 0:
                chg_cell.font = RED_FONT

            pct_cell = ws.cell(row=r, column=8, value=pct_change if pct_change is not None else "")
            if isinstance(pct_change, (int, float)):
                pct_cell.number_format = "0.0%"
                if pct_change > 0:
                    pct_cell.font = GREEN_FONT
                elif pct_change < 0:
                    pct_cell.font = RED_FONT

            ws.cell(row=r, column=9, value=status)

            # Row fill based on status
            if status == "SURGE":
                row_fill = green_fill
            elif status == "DECLINE":
                row_fill = light_red_fill
            elif status == "NEW":
                row_fill = light_yellow_fill
            elif status == "LOST":
                row_fill = light_gray_fill
            else:
                row_fill = zebra_fill if (i % 2 == 1) else None

            for col_idx in range(1, 10):
                cell = ws.cell(row=r, column=col_idx)
                cell.border = thin_border
                if row_fill:
                    cell.fill = row_fill
                if cell.font == Font():
                    cell.font = calibri(11)
                cell.alignment = Alignment(
                    horizontal="center" if col_idx != 2 else "left",
                    vertical="center"
                )
    else:
        # No comparison data available
        ws.merge_cells("A4:I4")
        cell = ws["A4"]
        cell.value = "No previous report found. Dealer metrics comparison will be available next month."
        cell.font = calibri(12, italic=True, color=GRAY)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 5
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 10

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 9: Vehicle Inventory
    # ══════════════════════════════════════════════════════════════════════
    log("  Sheet 9: Vehicle Inventory...")
    ws = wb.create_sheet("Vehicle Inventory")

    ws.merge_cells("A1:M1")
    cell = ws["A1"]
    cell.value = "Individual Vehicle Inventory"
    cell.font = calibri(16, bold=True, color=NAVY)
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:M2")
    cell = ws["A2"]
    cell.value = f"{total_vehicles:,} vehicles | {date_display} scrape"
    cell.font = calibri(11, italic=True, color=GRAY)
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 8

    set_header_row(ws, 4, ["VIN", "Dealer Name", "City", "State", "Brand", "Model",
                            "Body Type", "Body Builder", "Price", "Condition",
                            "Transmission", "Fuel Type", "Color"])
    ws.row_dimensions[4].height = 28

    vehicles_sorted = sorted(vehicles, key=lambda v: (v['dealer_name'], v['brand'], v['model']))

    for i, v in enumerate(vehicles_sorted):
        row = 5 + i
        ws.row_dimensions[row].height = 18
        fill = zebra_fill if i % 2 == 0 else None

        price_val = v['price_num'] if v['price_num'] > 0 else None

        vals = [v['vin'], v['dealer_name'], v['city'], v['state'], v['brand'],
                v['model'], v['body_type'], v['body_builder'], price_val,
                v['condition'], v['transmission'], v['fuel_type'], v['color']]
        fmts = [None, None, None, None, None, None, None, None, "$#,##0",
                None, None, None, None]
        aligns_row = [left_center, left_center, left_center, center, left_center,
                      left_center, left_center, left_center, center, center,
                      left_center, left_center, left_center]

        for j, (val, fmt, al) in enumerate(zip(vals, fmts, aligns_row)):
            cell = ws.cell(row=row, column=1 + j, value=val)
            cell.font = calibri(9)
            cell.border = thin_border
            cell.alignment = al
            if fmt and val is not None:
                cell.number_format = fmt
            if fill:
                cell.fill = fill

        if (i + 1) % 2000 == 0:
            log(f"    ... {i + 1:,} vehicles written")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 6
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 22
    ws.column_dimensions["I"].width = 12
    ws.column_dimensions["J"].width = 10
    ws.column_dimensions["K"].width = 22
    ws.column_dimensions["L"].width = 12
    ws.column_dimensions["M"].width = 16

    # ── Tab Colors ────────────────────────────────────────────────────────
    wb["Executive Summary"].sheet_properties.tabColor = "001F4E79"
    wb["Brand Summary"].sheet_properties.tabColor = "002E75B6"
    wb["Body Type Summary"].sheet_properties.tabColor = "002E75B6"
    wb["All Dealers"].sheet_properties.tabColor = "002E8B57"
    wb["Dealer Brand Matrix"].sheet_properties.tabColor = "002E75B6"
    wb["Dealer Body Type Matrix"].sheet_properties.tabColor = "002E75B6"
    wb["Smyrna Product Analysis"].sheet_properties.tabColor = "008B4513"
    wb["Dealer Metrics"].sheet_properties.tabColor = "00E67E22"
    wb["Vehicle Inventory"].sheet_properties.tabColor = "009B59B6"

    # ── Freeze Panes ─────────────────────────────────────────────────────
    wb["Executive Summary"].freeze_panes = "A5"
    wb["Brand Summary"].freeze_panes = "A5"
    wb["Body Type Summary"].freeze_panes = "A5"
    wb["All Dealers"].freeze_panes = "A5"
    wb["Dealer Brand Matrix"].freeze_panes = "E5"
    wb["Dealer Body Type Matrix"].freeze_panes = "E5"
    wb["Smyrna Product Analysis"].freeze_panes = "A5"
    wb["Dealer Metrics"].freeze_panes = "A8"
    wb["Vehicle Inventory"].freeze_panes = "A5"

    # ── Save ──────────────────────────────────────────────────────────────
    log(f"  Saving report to {report_path}...")
    wb.save(report_path)
    log(f"  Report saved successfully!")

    return report_path, total_vehicles, total_dealers, total_brands, total_body_types, total_smyrna, smyrna_dealers


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 5: VALIDATION
# ═══════════════════════════════════════════════════════════════════════════════

def run_validation(inventory_path, smyrna_vins, diff_stats, report_stats):
    """Step 5: Print validation summary and flag anomalies."""
    log_section("STEP 5: VALIDATION SUMMARY")

    report_path, total_vehicles, total_dealers, total_brands, total_body_types, \
        total_smyrna, smyrna_dealer_count = report_stats

    anomalies = []

    # Check vehicle count
    if total_vehicles < 5000:
        anomalies.append(f"LOW VEHICLE COUNT: {total_vehicles:,} (expected 10,000+)")
    elif total_vehicles < 10000:
        anomalies.append(f"BELOW AVERAGE VEHICLE COUNT: {total_vehicles:,} (typical ~13,000)")

    # Check dealer count
    if total_dealers < 200:
        anomalies.append(f"LOW DEALER COUNT: {total_dealers} (expected 500+)")
    elif total_dealers < 400:
        anomalies.append(f"BELOW AVERAGE DEALER COUNT: {total_dealers} (typical ~588)")

    # Check brand count
    if total_brands < 10:
        anomalies.append(f"MISSING BRANDS: only {total_brands} found (expected 13)")

    # Check body type count
    if total_body_types < 20:
        anomalies.append(f"MISSING BODY TYPES: only {total_body_types} found (expected 25)")

    # Check Smyrna count
    if total_smyrna == 0:
        anomalies.append("NO SMYRNA VEHICLES FOUND (WTS scrape may have failed)")
    elif total_smyrna < 30:
        anomalies.append(f"LOW SMYRNA COUNT: {total_smyrna} (expected ~75)")

    # Check diff stats
    if diff_stats:
        if diff_stats['new'] > total_vehicles * 0.5:
            anomalies.append(f"HIGH NEW VEHICLE RATE: {diff_stats['new']:,} "
                           f"({diff_stats['new']/total_vehicles*100:.0f}% of inventory)")
        if diff_stats['sold'] > diff_stats['previous_count'] * 0.5:
            anomalies.append(f"HIGH SOLD RATE: {diff_stats['sold']:,} "
                           f"({diff_stats['sold']/diff_stats['previous_count']*100:.0f}% of previous)")

    # Check for VIN-less dealers (dealers with missing names)
    inv = load_inventory_csv(inventory_path)
    missing_dealer = sum(1 for v in inv.values() if not v.get('dealer_name', '').strip())
    if missing_dealer > 0:
        pct = missing_dealer / len(inv) * 100
        anomalies.append(f"VEHICLES MISSING DEALER NAME: {missing_dealer} ({pct:.1f}%)")

    # Print summary
    log("")
    log(f"  Total Vehicles:     {total_vehicles:,}")
    log(f"  Total Dealers:      {total_dealers}")
    log(f"  Brands:             {total_brands}")
    log(f"  Body Types:         {total_body_types}")
    log(f"  Smyrna Products:    {total_smyrna} across {smyrna_dealer_count} dealers")
    log(f"  Smyrna WTS VINs:    {len(smyrna_vins)}")
    log("")

    if diff_stats:
        log(f"  vs Previous ({diff_stats['previous_file']}):")
        log(f"    New vehicles:     {diff_stats['new']:,}")
        log(f"    Sold vehicles:    {diff_stats['sold']:,}")
        log(f"    Price changes:    {diff_stats['price_changes']:,}")
        log(f"    Retained:         {diff_stats['retained']:,}")
        log("")

    log(f"  Report: {report_path}")
    log("")

    if anomalies:
        log("  *** ANOMALIES DETECTED ***")
        for a in anomalies:
            log(f"    ! {a}")
    else:
        log("  No anomalies detected. All checks passed.")

    log("")
    log("=" * 70)
    log("MONTHLY SCRAPE COMPLETE")
    log("=" * 70)

    return anomalies


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description='Comvoy Monthly Scrape -- Full end-to-end inventory intelligence pipeline'
    )
    parser.add_argument('--resume', action='store_true',
                        help='Resume an interrupted Comvoy scrape')
    parser.add_argument('--skip-scrape', action='store_true',
                        help='Skip scraping, use existing CSV for today (report-only mode)')
    parser.add_argument('--skip-smyrna', action='store_true',
                        help='Skip Smyrna WTS scrape, use most recent smyrna CSV')
    parser.add_argument('--old-report',
                        help='Path to previous month Excel report for dealer metrics comparison')
    parser.add_argument('--date', help='Override date (YYYY-MM-DD format)')
    args = parser.parse_args()

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    date_str = args.date or datetime.now().strftime('%Y-%m-%d')
    inventory_path = os.path.join(OUTPUT_DIR, f'inventory_{date_str}.csv')
    smyrna_path = os.path.join(OUTPUT_DIR, f'smyrna_inventory_{date_str}.csv')
    log_path = os.path.join(OUTPUT_DIR, f'monthly_scrape_log_{date_str}.txt')

    global LOG_FILE
    LOG_FILE = open(log_path, 'a', encoding='utf-8')

    start_time = time.time()

    log_section(f"COMVOY MONTHLY SCRAPE -- {date_str}")
    log(f"Output directory: {os.path.abspath(OUTPUT_DIR)}")

    # ── Step 1: Comvoy Inventory ──────────────────────────────────────────
    if args.skip_scrape:
        if os.path.exists(inventory_path):
            log(f"Skipping Comvoy scrape -- using existing {os.path.basename(inventory_path)}")
        else:
            log(f"ERROR: --skip-scrape but {inventory_path} does not exist!")
            sys.exit(1)
    else:
        run_comvoy_scrape(date_str, inventory_path, resume=args.resume)

    # ── Step 2: Smyrna WTS ───────────────────────────────────────────────
    smyrna_vins = set()
    smyrna_vehicles = []

    if args.skip_smyrna:
        # Find most recent Smyrna CSV
        smyrna_candidates = []
        for fname in os.listdir(OUTPUT_DIR):
            if fname.startswith('smyrna_inventory') and fname.endswith('.csv'):
                smyrna_candidates.append(os.path.join(OUTPUT_DIR, fname))
        if smyrna_candidates:
            smyrna_candidates.sort(key=os.path.getmtime, reverse=True)
            latest_smyrna = smyrna_candidates[0]
            log(f"Skipping Smyrna scrape -- using {os.path.basename(latest_smyrna)}")
            with open(latest_smyrna, encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    vin = row.get('vin', '').strip()
                    if vin:
                        smyrna_vins.add(vin)
                        smyrna_vehicles.append(row)
            log(f"  Loaded {len(smyrna_vins)} Smyrna VINs")
        else:
            log("WARNING: No Smyrna CSV found, Smyrna analysis will use body_builder field only")
    else:
        smyrna_vins, smyrna_vehicles = run_smyrna_scrape(date_str, smyrna_path)

    # ── Step 3: Diff ─────────────────────────────────────────────────────
    diff_stats = run_diff(date_str, inventory_path)

    # ── Step 4: Excel Report ─────────────────────────────────────────────
    old_report = args.old_report
    # If not provided, try known location
    if not old_report:
        default_old = os.path.join(os.path.expanduser('~'), 'Documents',
                                   'Comvoy_Multi_Brand_Report.xlsx')
        if os.path.exists(default_old):
            old_report = default_old

    report_stats = run_report(date_str, inventory_path, smyrna_vins,
                              smyrna_vehicles, diff_stats, old_report)

    # ── Step 5: Validation ───────────────────────────────────────────────
    anomalies = run_validation(inventory_path, smyrna_vins, diff_stats, report_stats)

    elapsed = time.time() - start_time
    log(f"Total runtime: {elapsed/60:.1f} minutes")

    LOG_FILE.close()

    # Print final files
    print(f"\n{'='*70}")
    print(f"OUTPUT FILES:")
    print(f"{'='*70}")
    print(f"  Inventory CSV:   {inventory_path}")
    print(f"  Smyrna CSV:      {smyrna_path}")
    if diff_stats:
        print(f"  New vehicles:    {os.path.join(OUTPUT_DIR, f'new_vehicles_{date_str}.csv')}")
        print(f"  Sold vehicles:   {os.path.join(OUTPUT_DIR, f'sold_vehicles_{date_str}.csv')}")
        print(f"  Price changes:   {os.path.join(OUTPUT_DIR, f'price_changes_{date_str}.csv')}")
    print(f"  Excel report:    {report_stats[0]}")
    print(f"  Log:             {log_path}")
    print(f"{'='*70}")

    if anomalies:
        print(f"\n  *** {len(anomalies)} ANOMALIES -- review log ***")
        sys.exit(2)
    else:
        print(f"\n  All checks passed.")


if __name__ == '__main__':
    main()
