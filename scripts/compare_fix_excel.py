import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter
from copy import copy
import os

ORIGINAL = r"C:\Users\motle\Documents\Comvoy_Multi_Brand_Report.xlsx"
NEW_FILE = r"C:\Users\motle\claude-code\comvoy\scrape_output\Comvoy_Multi_Brand_Report_2026-03-12.xlsx"

print("=" * 80)
print("STEP 1: ORIGINAL REPORT PROPERTIES")
print("=" * 80)

wb_orig = openpyxl.load_workbook(ORIGINAL)
orig_info = {}

for name in wb_orig.sheetnames:
    ws = wb_orig[name]
    props = ws.sheet_properties
    tab_color = props.tabColor
    print(f"\n--- Sheet: {name} ---")
    print(f"  Tab color: {tab_color}")
    if tab_color:
        print(f"    rgb={tab_color.rgb}, theme={tab_color.theme}, tint={tab_color.tint}, type={tab_color.type}")
    print(f"  Freeze panes: {ws.freeze_panes}")

    # Page setup (safely)
    try:
        ps = ws.page_setup
        print(f"  Page setup: orientation={ps.orientation}, paperSize={ps.paperSize}")
        print(f"    fitToWidth={ps.fitToWidth}, fitToHeight={ps.fitToHeight}")
    except Exception as e:
        print(f"  Page setup: error reading - {e}")

    # Sheet format
    print(f"  Sheet format: defaultRowHeight={ws.sheet_format.defaultRowHeight}")

    # Collect row heights for first 15 rows
    row_heights = {}
    for r in range(1, 16):
        rd = ws.row_dimensions.get(r)
        if rd and rd.height is not None:
            row_heights[r] = rd.height
    print(f"  Row heights (first 15): {row_heights}")

    # Column widths
    col_widths = {}
    for col_letter, cd in ws.column_dimensions.items():
        if cd.width is not None and cd.width != 8.0:
            col_widths[col_letter] = cd.width
    print(f"  Column widths: {col_widths}")

    # Check formatting of first few rows
    fmt_info = {}
    for r in range(1, 6):
        for c in range(1, min(ws.max_column + 1, 15)):
            cell = ws.cell(r, c)
            if cell.value is not None or (cell.font and cell.font.bold):
                key = f"R{r}C{c}"
                fc = cell.font.color
                fc_rgb = None
                if fc:
                    if fc.rgb and str(fc.rgb) != "00000000":
                        fc_rgb = str(fc.rgb)
                    elif fc.theme is not None:
                        fc_rgb = f"theme={fc.theme}"
                fl = cell.fill
                fl_rgb = None
                if fl and fl.fgColor:
                    if fl.fgColor.rgb and str(fl.fgColor.rgb) != "00000000":
                        fl_rgb = str(fl.fgColor.rgb)
                    elif fl.fgColor.theme is not None:
                        fl_rgb = f"theme={fl.fgColor.theme}"
                fmt_info[key] = {
                    "value": str(cell.value)[:40] if cell.value else None,
                    "font_name": cell.font.name,
                    "font_size": cell.font.size,
                    "font_bold": cell.font.bold,
                    "font_italic": cell.font.italic,
                    "font_color": fc_rgb,
                    "fill_color": fl_rgb,
                    "fill_type": cell.fill.fill_type,
                }
    print(f"  Cell formatting (first 5 rows):")
    for k, v in fmt_info.items():
        if v["value"] or v["font_bold"]:
            print(f"    {k}: val={v['value']}, size={v['font_size']}, bold={v['font_bold']}, italic={v['font_italic']}, font_color={v['font_color']}, fill={v['fill_color']}, fill_type={v['fill_type']}")

    orig_info[name] = {
        "tab_color": tab_color,
        "freeze": ws.freeze_panes,
        "row_heights": row_heights,
        "col_widths": col_widths,
        "fmt": fmt_info,
    }

print("\n\n" + "=" * 80)
print("STEP 2: NEW REPORT PROPERTIES")
print("=" * 80)

wb_new = openpyxl.load_workbook(NEW_FILE)
new_info = {}

for name in wb_new.sheetnames:
    ws = wb_new[name]
    props = ws.sheet_properties
    tab_color = props.tabColor
    print(f"\n--- Sheet: {name} ---")
    print(f"  Tab color: {tab_color}")
    if tab_color:
        print(f"    rgb={tab_color.rgb}, theme={tab_color.theme}, tint={tab_color.tint}, type={tab_color.type}")
    print(f"  Freeze panes: {ws.freeze_panes}")

    row_heights = {}
    for r in range(1, 16):
        rd = ws.row_dimensions.get(r)
        if rd and rd.height is not None:
            row_heights[r] = rd.height
    print(f"  Row heights (first 15): {row_heights}")

    col_widths = {}
    for col_letter, cd in ws.column_dimensions.items():
        if cd.width is not None and cd.width != 8.0:
            col_widths[col_letter] = cd.width
    print(f"  Column widths: {col_widths}")

    fmt_info = {}
    for r in range(1, 6):
        for c in range(1, min(ws.max_column + 1, 15)):
            cell = ws.cell(r, c)
            if cell.value is not None or (cell.font and cell.font.bold):
                key = f"R{r}C{c}"
                fc = cell.font.color
                fc_rgb = None
                if fc:
                    if fc.rgb and str(fc.rgb) != "00000000":
                        fc_rgb = str(fc.rgb)
                    elif fc.theme is not None:
                        fc_rgb = f"theme={fc.theme}"
                fl = cell.fill
                fl_rgb = None
                if fl and fl.fgColor:
                    if fl.fgColor.rgb and str(fl.fgColor.rgb) != "00000000":
                        fl_rgb = str(fl.fgColor.rgb)
                    elif fl.fgColor.theme is not None:
                        fl_rgb = f"theme={fl.fgColor.theme}"
                fmt_info[key] = {
                    "value": str(cell.value)[:40] if cell.value else None,
                    "font_name": cell.font.name,
                    "font_size": cell.font.size,
                    "font_bold": cell.font.bold,
                    "font_italic": cell.font.italic,
                    "font_color": fc_rgb,
                    "fill_color": fl_rgb,
                    "fill_type": cell.fill.fill_type,
                }
    print(f"  Cell formatting (first 5 rows):")
    for k, v in fmt_info.items():
        if v["value"] or v["font_bold"]:
            print(f"    {k}: val={v['value']}, size={v['font_size']}, bold={v['font_bold']}, italic={v['font_italic']}, font_color={v['font_color']}, fill={v['fill_color']}, fill_type={v['fill_type']}")

    new_info[name] = {
        "tab_color": tab_color,
        "freeze": ws.freeze_panes,
        "row_heights": row_heights,
        "col_widths": col_widths,
        "fmt": fmt_info,
    }

wb_orig.close()

print("\n\n" + "=" * 80)
print("STEP 3: DIFFERENCES")
print("=" * 80)

common_sheets = [s for s in wb_orig.sheetnames if s in wb_new.sheetnames]
new_only = [s for s in wb_new.sheetnames if s not in wb_orig.sheetnames]
orig_only = [s for s in wb_orig.sheetnames if s not in wb_new.sheetnames]

if orig_only:
    print(f"\nSheets only in original: {orig_only}")
if new_only:
    print(f"\nSheets only in new report: {new_only}")

for name in common_sheets:
    diffs = []
    o = orig_info[name]
    n = new_info[name]

    # Tab color
    def tc_str(tc):
        if tc is None:
            return "None"
        if tc.rgb:
            return str(tc.rgb)
        if tc.theme is not None:
            return f"theme={tc.theme}"
        return str(tc)

    o_tc = tc_str(o["tab_color"])
    n_tc = tc_str(n["tab_color"])
    if o_tc != n_tc:
        diffs.append(f"  Tab color: orig={o_tc} vs new={n_tc}")

    # Freeze panes
    if o["freeze"] != n["freeze"]:
        diffs.append(f"  Freeze panes: orig={o['freeze']} vs new={n['freeze']}")

    # Row heights
    all_rows = set(list(o["row_heights"].keys()) + list(n["row_heights"].keys()))
    for r in sorted(all_rows):
        oh = o["row_heights"].get(r)
        nh = n["row_heights"].get(r)
        if oh != nh:
            diffs.append(f"  Row {r} height: orig={oh} vs new={nh}")

    # Column widths
    all_cols = set(list(o["col_widths"].keys()) + list(n["col_widths"].keys()))
    for c in sorted(all_cols):
        ow = o["col_widths"].get(c)
        nw = n["col_widths"].get(c)
        if ow != nw:
            diffs.append(f"  Col {c} width: orig={ow} vs new={nw}")

    # Cell formatting
    all_cells = set(list(o["fmt"].keys()) + list(n["fmt"].keys()))
    for cell_key in sorted(all_cells):
        of = o["fmt"].get(cell_key, {})
        nf = n["fmt"].get(cell_key, {})
        for prop in ["font_size", "font_bold", "font_italic", "font_color", "fill_color", "fill_type"]:
            ov = of.get(prop)
            nv = nf.get(prop)
            if ov != nv:
                diffs.append(f"  {cell_key} {prop}: orig={ov} vs new={nv}")

    if diffs:
        print(f"\n--- {name} ---")
        for d in diffs:
            print(d)
    else:
        print(f"\n--- {name} --- NO DIFFERENCES")


print("\n\n" + "=" * 80)
print("STEP 4: APPLYING FIXES")
print("=" * 80)

changes = []

# Re-open original to get tab colors and formatting
wb_orig = openpyxl.load_workbook(ORIGINAL)

# Apply tab colors from original to matching sheets
for name in common_sheets:
    ws_orig = wb_orig[name]
    ws_new = wb_new[name]

    orig_tc = ws_orig.sheet_properties.tabColor
    if orig_tc:
        ws_new.sheet_properties.tabColor = copy(orig_tc)
        changes.append(f"Set tab color for '{name}' to {orig_tc.rgb}")

    # Fix freeze panes
    orig_freeze = ws_orig.freeze_panes
    new_freeze = ws_new.freeze_panes
    if orig_freeze != new_freeze:
        ws_new.freeze_panes = orig_freeze
        changes.append(f"Set freeze panes for '{name}' to {orig_freeze}")

    # Fix row heights for ALL rows
    for r in range(1, max(ws_orig.max_row or 1, ws_new.max_row or 1) + 1):
        rd_orig = ws_orig.row_dimensions.get(r)
        if rd_orig and rd_orig.height is not None:
            rd_new = ws_new.row_dimensions.get(r)
            if not rd_new or rd_new.height != rd_orig.height:
                ws_new.row_dimensions[r].height = rd_orig.height
                if r <= 10:
                    changes.append(f"Set row {r} height for '{name}' to {rd_orig.height}")

    # Fix column widths
    for col_letter, cd_orig in ws_orig.column_dimensions.items():
        if cd_orig.width is not None:
            cd_new = ws_new.column_dimensions.get(col_letter)
            if not cd_new or cd_new.width != cd_orig.width:
                ws_new.column_dimensions[col_letter].width = cd_orig.width
                changes.append(f"Set col {col_letter} width for '{name}' to {cd_orig.width}")

    # Fix cell formatting for header area (first 5 rows)
    for r in range(1, 6):
        for c in range(1, min((ws_orig.max_column or 1) + 1, 30)):
            cell_orig = ws_orig.cell(r, c)
            cell_new = ws_new.cell(r, c)

            # Compare fonts
            orig_font = cell_orig.font
            new_font = cell_new.font
            font_diff = (
                orig_font.size != new_font.size or
                orig_font.bold != new_font.bold or
                orig_font.italic != new_font.italic or
                orig_font.name != new_font.name
            )
            if font_diff:
                cell_new.font = copy(orig_font)
                if cell_new.value:
                    changes.append(f"Fixed font for '{name}' R{r}C{c} (size={orig_font.size}, bold={orig_font.bold})")

            # Compare fills
            orig_fill = cell_orig.fill
            new_fill = cell_new.fill
            if orig_fill.fill_type != new_fill.fill_type:
                cell_new.fill = copy(orig_fill)
                if cell_new.value:
                    changes.append(f"Fixed fill for '{name}' R{r}C{c}")

    # Copy page setup (safely)
    try:
        ws_new.page_setup.orientation = ws_orig.page_setup.orientation
        ws_new.page_setup.paperSize = ws_orig.page_setup.paperSize
        ws_new.page_setup.fitToWidth = ws_orig.page_setup.fitToWidth
        ws_new.page_setup.fitToHeight = ws_orig.page_setup.fitToHeight
    except Exception:
        pass

# Pick tab colors for new sheets based on the palette
print("\nExisting tab colors palette:")
palette = {}
for name in wb_orig.sheetnames:
    tc = wb_orig[name].sheet_properties.tabColor
    if tc and tc.rgb:
        palette[name] = tc.rgb
        print(f"  {name}: {tc.rgb}")

# Assign complementary colors to new sheets
color_assignments = {
    "Dealer Metrics": "00E67E22",      # Orange - analytics/metrics feel
    "Vehicle Inventory": "009B59B6",    # Purple - inventory/catalog feel
}

for name in new_only:
    ws = wb_new[name]
    if name in color_assignments:
        ws.sheet_properties.tabColor = Color(rgb=color_assignments[name])
        changes.append(f"Set tab color for NEW sheet '{name}' to {color_assignments[name]}")

    # Set freeze panes if not set (freeze below header row)
    if not ws.freeze_panes:
        # Find the header row by looking for bold cells followed by data
        for r in range(1, 10):
            cell = ws.cell(r, 1)
            if cell.value and cell.font and cell.font.bold:
                next_cell = ws.cell(r + 1, 1)
                if next_cell.value and not (next_cell.font and next_cell.font.bold):
                    ws.freeze_panes = f"A{r + 1}"
                    changes.append(f"Set freeze panes for NEW sheet '{name}' to A{r + 1}")
                    break

# Save
wb_new.save(NEW_FILE)
wb_orig.close()
wb_new.close()

print("\n" + "=" * 80)
print("CHANGES APPLIED")
print("=" * 80)
for c in changes:
    print(f"  - {c}")
print(f"\nTotal changes: {len(changes)}")
print(f"File saved: {NEW_FILE}")
